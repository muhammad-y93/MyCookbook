from pathlib import Path
import openpyxl as xl
from openpyxl.chart import PieChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']
    # cell = sheet['a1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        new_price = cell.value * 0.9
        new_price_cell = sheet.cell(row, 4)
        new_price_cell.value = new_price

    values = Reference(sheet, min_row = 2, max_row = sheet.max_row, min_col = 4, max_col = 4)
    chart = PieChart()
    sheet.add_chart(chart, 'e2')
    wb.save('filename')


# mkdir()   =>  To make directory
# rmdir()   =>  To remove directory
# glob()    =>  To list all the items in a directory

# path = Path('')
# for file in path.glob('*'):
#     print(file)

